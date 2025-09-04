import openpyxl
from openpyxl_image_loader import SheetImageLoader
import io
import base64
from PIL import Image
import logging
import os
from pathlib import Path
import urllib.parse
import pickle
import json
import re
from dotenv import load_dotenv
import requests
from openpyxl.utils import get_column_letter
import io
from openpyxl_image_loader import SheetImageLoader
from PIL import Image as PILImage

load_dotenv()

GENERAL_PURPOSE_API_URL = "https://vib.cagent.cmcts.ai/api/v1/prediction/33059c88-4ae9-407d-920a-6efc9404c814"
SEND_RESPONSE_API_URL = os.getenv("SEND_RESPONSE_API_URL")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def create_response(file_name, data):
    """
    Tạo response dạng base64 cho các file theo cấu trúc yêu cầu
    
    Args:
        data (dict): Dữ liệu JSON chứa thông tin các process và steps
        
    Returns:
        dict: Response chứa base64 của file tổng quan và các file cụ thể
    """
    response = {"file_name": file_name, "data": []}
    # Tạo file tổng quan
    sheet_tong_quan = data["processes"]["sheet_tong_quan"]
    content = f"#{sheet_tong_quan['title']}\n"
    for step in sheet_tong_quan["steps"]:
        content += f"\n## {step['title']}\n{step['content']}\n"
    response["data"].append({
        "file_path": os.path.join("vib-sample", "Pilot", "general", Path(file_name).stem, "tong_quan.txt"),
        "content": content,
    })
    
    # Tạo danh sách các file cụ thể
    sheet_cu_the = data["processes"]["sheet_cu_the"]
    for index, sheet in enumerate(sheet_cu_the):
        # Tạo nội dung metadata nếu có
        if sheet["metadata"]:
            content = "# Metadata\n"
            metadata = sheet["metadata"]
            for md in metadata:
                content += f"## Nhóm quy trình / Quy trình lớn: {md['general_process']}\n"
                content += f"### Quy trình cụ thể/ Quy trình con: {md['sub_process']}\n"
        else:
            content = ""
            
        # Thêm nội dung chính của sheet
        content += f"# {sheet['title']}\n"
        for step in sheet["steps"]:
            content += f"{step['id']}. {step['title']}\n{step['content']}\n"
            
        response["data"].append({
            "file_path": os.path.join("vib-sample", "Pilot","specific", Path(file_name).stem, f"{index+1}.txt"),
            "content": content,
        })
    

    return response

def extract_json(response_text):
    pattern = r"```json\s*([\[{].*?[\]}])\s*```"
    match = re.search(pattern, response_text, re.DOTALL)
    if match:
        return json.loads(match.group(1))
    return None

def encode_s3_path(path):
    """
    Encode S3 path: encode spaces and forward slashes, keep other characters unchanged
    
    Args:
        path (str): Original path
        
    Returns:
        str: Encoded path
    """
    # Encode the entire path, replacing spaces with %20 and forward slashes with %2F
    return path.replace(' ', '%20').replace('/', '%2F')

def format_markdown_cell(text):
    if not text:
        return ""
    text = text.replace("|", "\\|")
    text = text.replace("\n", " ")
    return text.strip()

def process_excel_to_markdown(file_name, file_content):
    """
    Process Excel file content and convert each sheet to markdown format with base64 encoded images
    
    Args:
        file_content (bytes): Content of Excel file
        
    Returns:
        list: List of dictionaries containing sheet name, markdown content and images
    """
    logger.info("Bắt đầu xử lý nội dung file Excel")
    results = []
    
    try:
        logger.info("Đang tải workbook...")
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        logger.info(f"Đã tải workbook thành công. Số sheet: {len(wb.sheetnames)}")
        
        
        for sheet_name in wb.sheetnames:
            logger.info(f"Đang xử lý sheet: {sheet_name}")
            sheet = wb[sheet_name]
            image_loader = SheetImageLoader(sheet)
            images = []
            
            image_map = {}
            for image in getattr(sheet, '_images', []):
                anchor = image.anchor                 
                from_col = anchor._from.col
                from_row = anchor._from.row
                coord = f"{get_column_letter(from_col + 1)}{from_row + 1}"
                image_map[coord] = image

            rows = list(sheet.iter_rows(values_only=False))
            
            # Tìm cột cuối cùng có dữ liệu
            max_col = 0
            for row in rows:
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row[0].row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        max_col = max(max_col, col)
            
            logger.info(f"Sheet {sheet_name}: {len(rows)} hàng, {max_col} cột có dữ liệu")
            
            # Process header
            logger.info(f"Đang xử lý header của sheet {sheet_name}")
            header = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                header.append(format_markdown_cell(str(cell.value)) if cell.value is not None else "")
                
            all_rows = [header]
            
            # Process data rows
            logger.info(f"Đang xử lý dữ liệu của sheet {sheet_name}")
            for row_idx, row in enumerate(rows[1:], start=2):
                logger.debug(f"Đang xử lý hàng {row_idx}")
                row_data = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    coord = cell.coordinate
                    try:
                        text = str(cell.value) if cell.value else ""
                        if coord in image_map:
                            img = image_map[coord]
                            # img_buffer = io.BytesIO(img.ref)  # ref là bytes của ảnh
                            if isinstance(img.ref, bytes):
                                img_buffer = io.BytesIO(img.ref)
                            elif isinstance(img.ref, io.BytesIO):
                                img_buffer = img.ref
                            else:
                                logger.error(f"Không thể xác định kiểu ảnh ở ô {coord}: {type(img.ref)}")
                                continue

                            pil_img = PILImage.open(img_buffer)

                            img_out = io.BytesIO()
                            # Đọc bytes từ BytesIO và encode
                            
                            pil_img.save(img_out, format="PNG")
                            img_out.seek(0)
                            img_bytes = img_out.getvalue()
                            # print(type(img_bytes))
                            img_base64 = base64.b64encode(img_bytes).decode("utf-8")

                            s3_prefix = f"vib-images/imaghttps://vib.cagent.cmcts.ai/api/v1/prediction/33059c88-4ae9-407d-920a-6efc9404c814es/{file_name}/sheet_{sheet_name}_{coord}.png"
                            # encoded_prefix = encode_s3_path(s3_prefix)
                            s3_url = f"/s3-explorer/api/download?path={s3_prefix}"
                            images.append({
                                "coordinate": coord,
                                "data": img_base64,
                                "content_type": "image/png",
                                "prefix": s3_prefix
                            })

                            # ✅ Thay vì chèn ảnh markdown, chỉ chèn dòng mô tả tham khảo
                            text += f" ![Image]({s3_url})"
                        # if image_loader.image_in(coord):
                        #     logger.debug(f"Phát hiện ảnh ở ô {coord}")
                        #     try:
                        #         img = image_loader.get(coord)
                        #         img_buffer = io.BytesIO()
                        #         img.save(img_buffer, format='PNG')
                        #         img_buffer.seek(0)
                                
                        #         # Convert image to base64
                        #         image_data = img_buffer.getvalue()
                        #         img_base64 = base64.b64encode(image_data).decode('utf-8')
                                
                        #         # Create image prefix and S3 URL
                        #         file_name = f"{coord}.png"
                        #         s3_prefix = f"vib-images/{excel_name}/{sheet_name}/{file_name}"
                        #         encoded_prefix = encode_s3_path(s3_prefix)
                        #         s3_url = f"/s3-explorer/api/download?path={encoded_prefix}"
                                
                        #         images.append({
                        #             "coordinate": coord,
                        #             "data": img_base64,
                        #             "content_type": "image/png",
                        #             "prefix": s3_prefix
                        #         })
                                
                        #         # Add image reference to text using S3 URL
                        #         if "[ICON]" in text:
                        #             text = text.replace("[ICON]", f"![{coord}]({s3_url})")
                        #         else:
                        #             text += f" ![Image]({s3_url})"
                                    
                        #         img_buffer.close()
                        #         logger.debug(f"Đã xử lý ảnh ở ô {coord} thành công")
                                    
                        #     except Exception as img_e:
                        #         logger.error(f"Lỗi xử lý ảnh ở ô {coord}: {img_e}")
                                
                        row_data.append(format_markdown_cell(text))
                        
                    except Exception as e:
                        logger.error(f"Lỗi xử lý ô {coord}: {e}")
                        row_data.append("")
                all_rows.append(row_data)
                
            # Generate markdown content
            logger.info(f"Đang tạo markdown cho sheet {sheet_name}")
            markdown_content = []
            markdown_content.append("| " + " | ".join(all_rows[0]) + " |")
            markdown_content.append("|" + "|".join(["---"] * len(all_rows[0])) + "|")
            for row in all_rows[1:]:
                markdown_content.append("| " + " | ".join(row) + " |")
                
            results.append({
                "sheet_name": sheet_name,
                "content": "\n".join(markdown_content),
                "images": images
            })
            logger.info(f"Đã xử lý xong sheet {sheet_name}. Số ảnh: {len(images)}")
            with open("results.pkl", "wb") as f:
                pickle.dump(results, f)
        wb.close()
        logger.info(f"Đã xử lý xong file Excel. Tổng số sheet: {len(results)}")

        return results   
        
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file Excel: {e}")
        raise

def query(payload):
    import requests
    response = requests.post(GENERAL_PURPOSE_API_URL, json=payload)
    try:
        return response.json()
    except Exception as e:
        logger.error(f"Lỗi khi query API: {e}")
        return response

def api_call(file_name, file_content):
    results = process_excel_to_markdown(file_name, file_content)
    
    # Gộp nội dung các sheet vào một content duy nhất
    combined_content = []
    for result in results:
        combined_content.append(f"## Sheet {result['sheet_name']}:")
        combined_content.append(result["content"])
        combined_content.append("")  # Thêm dòng trống giữa các sheet
    
    merged_results = {
        "content": "\n".join(combined_content),
        "images": [img for result in results for img in result["images"]]
    }
    
    return merged_results
#     # Tìm sheet tổng hợp
#     detect_summary_sheet_system_prompt = """
# Bạn là một trợ lý phân tích nghiệp vụ nội bộ có khả năng đọc hiểu các bảng dữ liệu nghiệp vụ từ các tài liệu hướng dẫn nội bộ, đặc biệt là hệ thống ACL.
# Nội dung tài liệu bảo gồm sheet tổng quan, và sheet các bước cụ thể trong một quy trình được tham chiếu từ sheet tổng quan

# ## Instructions following
# 1. Xác định đâu là sheet tổng quan
# - Có các cột tiêu đề như:
#     - Item
#     - Nhóm nghiệp vụ
#     - Nhóm công việc
#     - Quy trình ACL
#     - Bước
#     - Tên bước thực hiện
#     - Link hoặc tham chiếu
# - Có các mô tả chi tiết về các bước thực hiện
# - Có các liên kết giữa các quy trình

# 2. Trích xuất toàn bộ quy trình
# - Hãy trích xuất tất cả quy trình (hoặc mục lớn) và các bước chi tiết của quy trình (hoặc mục lớn) đó từ nội dung tài liệu đưa ra


# ## YÊU CẦU VỀ ĐẦU RA
# Đầu ra của phản hồi phải là một json schema như sau:
# ```json
# {{  
#     "file_title": "tiêu đề văn bản"
#     "processes": {{
#         "sheet_tong_quan": {{
#             "title": "STT. Tên quy trình (hoặc mục lớn)",
#             "steps": [
#                 {{
#                     "id": "STT quy trình con",
#                     "title": "Tên quy trình con",
#                     "content": "Đầy đủ các đề mục và nội dung từng đề mục trong bước (ví dụ: người thực hiện, nội dung, ...)"
#                 }},
#                 ...
#             ]
#         }},
#         "sheet_cu_the": [
#         {{
#             "title": "Tên bước",
#             "steps": [
#                 {{
#                     "id": "STT bước",
#                     "title": "Tên bước hoặc tóm tắt nội dung bước",
#                     "content": "Đầy đủ các đề mục và nội dung từng đề mục trong bước (ví dụ: người thực hiện, nội dung, ...)"
#                 }},
#             "metadata": [
#                 {{
#                     "general_process": "Tên quy trình lớn có tham chiếu đến bước cụ thể này",
#                     "sub_process": "Tên quy trình con có tham chiếu đến bước cụ thể này"
#                 }}
#             ]
#                 ...
#             ]
#         }}
#     ]
#     }}
# }}
# ```
# YÊU CẦU:
# 1. Trích xuất đầy đủ tất cả các quy trình
# 2. Mỗi bước phải có ID, tiêu đề và nội dung chi tiết
# 3. Đảm bảo tính chính xác của cấu trúc JSON
# 4. Nội dung trong từng bước phải dưới dạng text thuần, đầy đủ các mục, không sử dụng bảng.
# TÀI LIỆU ĐẦU VÀO:
# """.strip()
    
#     payload = {
#         "question": merged_results["content"],
#         "overrideConfig": {
#             "systemMessagePrompt": detect_summary_sheet_system_prompt
#         }
#     }

#     response = query(payload)
#     json_response = extract_json(response["text"])
    
#     # Tạo base64 response
#     response = create_response(file_name, json_response)

#     headers = {
#         'accept': '',
#         'Content-Type': 'application/json'
#     }

#     send_response = requests.post(SEND_RESPONSE_API_URL, headers=headers, data=json.dumps(response))
#     if send_response.status_code == 200:
#         if send_response.json()["success"] == True:
#             return {"status": "success"}
        
#     return {"status": "error", "error_code": send_response.status_code}
